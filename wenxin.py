# -*- coding: utf-8 -*-
"""
整合脚本：从Excel读取数据，提取夸克链接，直接上传图片到文心智能体
使用 uploadFileToBos 接口直接上传图片，无需经过百度BOS
"""
import re
import os
import sys
import json
import time
import logging
import requests
import yaml
from io import BytesIO
from openpyxl import load_workbook

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# 伪装浏览器请求头
DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Sec-Fetch-Dest": "image",
    "Sec-Fetch-Mode": "no-cors",
    "Sec-Fetch-Site": "cross-site",
    "Pragma": "no-cache",
    "Cache-Control": "no-cache"
}


def load_config(config_path="config.yaml"):
    """加载配置文件"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在: {config_path}")
    
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def setup_logging(config):
    """配置日志"""
    log_level = config.get('logging', {}).get('level', 'INFO')
    log_file = config.get('logging', {}).get('log_file', 'upload_direct.log')
    
    logging.basicConfig(
        level=getattr(logging, log_level),
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )


class QuarkAPI:
    """夸克网盘API"""
    BASE_URL = {
        "token": "https://drive-h.quark.cn/1/clouddrive/share/sharepage/token",
        "detail": "https://drive-m.quark.cn/1/clouddrive/share/sharepage/detail"
    }
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0",
            "Accept": "application/json, text/javascript, */*; q=0.01",
        })
    
    def extract_quark_url(self, text):
        """从文本中提取夸克网盘链接"""
        if not text or not isinstance(text, str):
            return None
        
        pattern = r"https?://pan\.quark\.cn/s/[\w\-]+"
        matches = re.findall(pattern, text)
        
        if matches:
            logging.info(f"提取到夸克链接: {matches[0]}")
            return matches[0]
        
        logging.warning("未找到夸克网盘链接")
        return None
    
    def get_id_from_url(self, url):
        """从夸克网盘分享链接中提取ID"""
        try:
            url = url.replace("https://pan.quark.cn/s/", "")
            pattern = r"(\w+)(#/list/share.*/(\w+))?"
            match = re.search(pattern, url)
            if match:
                pwd_id = match.group(1)
                pdir_fid = match.group(3) if match.group(2) else "0"
                return (pwd_id, pdir_fid)
            return None
        except Exception as e:
            logging.error("解析URL失败: " + str(e))
            return None

    def get_stoken(self, pwd_id):
        """获取分享token"""
        try:
            params = {"pr": "ucpro", "fr": "h5"}
            payload = {"pwd_id": pwd_id, "passcode": ""}
            response = self.session.post(
                self.BASE_URL["token"],
                json=payload,
                params=params
            ).json()
            
            if response.get("data"):
                return (True, response["data"]["stoken"])
            return (False, response.get("message", "未知错误"))
        except Exception as e:
            logging.error("获取stoken失败: " + str(e))
            return (False, str(e))

    def get_detail(self, pwd_id, stoken, pdir_fid):
        """获取文件详情"""
        try:
            params = {
                "pr": "ucpro",
                "fr": "pc",
                "pwd_id": pwd_id,
                "stoken": stoken[1],
                "pdir_fid": pdir_fid,
                "force": "0",
                "_page": 1,
                "_size": "500",
                "_fetch_banner": "0",
                "_fetch_share": "0",
                "_fetch_total": "1",
                "_sort": "file_type:asc,updated_at:desc",
            }
            
            response = self.session.get(
                self.BASE_URL["detail"],
                params=params
            ).json()
            
            if "data" not in response:
                logging.error("获取详情失败，响应: " + str(response))
                return None
                
            return response["data"]["list"]
        except Exception as e:
            logging.error("获取文件详情失败: " + str(e))
            return None

    def get_preview_urls(self, file_list):
        """获取预览URL列表"""
        real_images = []
        video_previews = []
        
        for item in file_list:
            format_type = item.get("format_type", "")
            file_name = item.get("file_name", "").lower()
            preview_url = item.get("preview_url")
            
            if preview_url:
                is_real_image = False
                
                if file_name.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp")):
                    is_real_image = True
                elif format_type.startswith("image/") and "video" not in file_name:
                    is_real_image = True
                
                if is_real_image:
                    real_images.append({
                        "file_name": item.get("file_name", ""),
                        "preview_url": preview_url,
                        "file_size": item.get("file_size"),
                        "updated_at": item.get("updated_at"),
                        "format_type": format_type
                    })
                else:
                    video_previews.append({
                        "file_name": item.get("file_name", ""),
                        "preview_url": preview_url,
                        "file_size": item.get("file_size"),
                        "updated_at": item.get("updated_at"),
                        "format_type": format_type
                    })
        
        if real_images:
            logging.info(f"找到 {len(real_images)} 个真实图片文件")
            return real_images
        else:
            logging.info(f"未找到真实图片，使用 {len(video_previews)} 个视频预览图")
            return video_previews

    def get_first_image_url(self, url):
        """从分享链接获取第一张图片的预览URL"""
        try:
            id_info = self.get_id_from_url(url)
            if not id_info:
                logging.error("无效的分享链接")
                return None, None
                
            pwd_id, pdir_fid = id_info
            logging.info("解析到ID: " + pwd_id + ", " + pdir_fid)
            
            stoken = self.get_stoken(pwd_id)
            if not stoken[0]:
                logging.error("获取token失败: " + stoken[1])
                return None, None
            
            detail = self.get_detail(pwd_id, stoken, pdir_fid)
            if not detail:
                logging.error("获取文件列表失败")
                return None, None
            
            if detail and len(detail) > 0:
                first_item = detail[0]
                logging.info(f"获取第一个项目的子内容: {first_item.get('file_name', 'unknown')}")
                
                sub_detail = self.get_detail(pwd_id, stoken, first_item["fid"])
                if sub_detail:
                    logging.info(f"成功获取子内容，包含 {len(sub_detail)} 个文件")
                    preview_urls = self.get_preview_urls(sub_detail)
                    if preview_urls and len(preview_urls) > 0:
                        logging.info(f"在子内容中找到 {len(preview_urls)} 个图片")
                        logging.info(f"使用第一个图片: {preview_urls[0]['file_name']}")
                        return preview_urls[0]["preview_url"], preview_urls[0]["file_name"]
                    else:
                        logging.warning("在子内容中未找到图片")
                else:
                    logging.warning("获取子内容失败")
            else:
                logging.warning("未获取到文件列表或文件列表为空")
            
            preview_urls = self.get_preview_urls(detail)
            if preview_urls and len(preview_urls) > 0:
                logging.info(f"在当前目录找到 {len(preview_urls)} 个图片文件")
                logging.info(f"使用第一个图片: {preview_urls[0]['file_name']}")
                return preview_urls[0]["preview_url"], preview_urls[0]["file_name"]
            
            logging.info("未找到图片")
            return None, None
            
        except Exception as e:
            logging.error("获取图片URL失败: " + str(e))
            return None, None


class AgentImageUploader:
    """文心智能体图片上传器 - 直接使用 uploadFileToBos 接口"""
    def __init__(self, config):
        agent_config = config.get('agent', {})
        self.cookie_file = agent_config.get('cookie_file')
        self.upload_url = "https://agents.baidu.com/lingjing/tools/uploadFileToBos"
        
        if not self.cookie_file or not os.path.exists(self.cookie_file):
            raise FileNotFoundError(f"Cookie文件不存在: {self.cookie_file}")
        
        self.cookie = self._load_cookie()
        self.session = requests.Session()
    
    def _load_cookie(self):
        """加载Cookie"""
        try:
            with open(self.cookie_file, 'r', encoding='utf-8') as f:
                content = f.read().strip().replace('`', '')
                return content
        except Exception as e:
            logging.error(f"读取Cookie失败: {e}")
            return None
    
    def _get_quark_specific_headers(self):
        """获取夸克网盘专用请求头"""
        return {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0",
            "Accept": "image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-US;q=0.7",
            "Origin": "https://pan.quark.cn",
            "Referer": "https://pan.quark.cn/",
            "Sec-Fetch-Dest": "image",
            "Sec-Fetch-Mode": "no-cors",
            "Sec-Fetch-Site": "cross-site",
            "Connection": "keep-alive"
        }
    
    def convert_image_format(self, image_data, target_format='JPEG'):
        """转换图片格式"""
        if not HAS_PIL:
            logging.warning("未安装PIL库，跳过格式转换")
            return image_data, '.webp'
        
        try:
            img = Image.open(BytesIO(image_data))
            
            if img.mode in ('RGBA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            output_buffer = BytesIO()
            img.save(output_buffer, format=target_format, quality=95)
            output_buffer.seek(0)
            
            if target_format == 'JPEG':
                ext = '.jpg'
            elif target_format == 'PNG':
                ext = '.png'
            else:
                ext = '.jpg'
            
            logging.info(f"图片格式已转换为 {target_format}")
            return output_buffer.getvalue(), ext
            
        except Exception as e:
            logging.warning(f"格式转换失败，使用原始格式: {str(e)}")
            return image_data, '.webp'
    
    def download_image(self, image_url):
        """从URL下载图片"""
        logging.info("正在下载图片: " + image_url)
        
        is_quark_image = "quark.cn" in image_url or "drive.quark.cn" in image_url
        
        if is_quark_image:
            logging.info("检测到夸克网盘图片，使用专用请求头")
            headers = self._get_quark_specific_headers()
        else:
            headers = DEFAULT_HEADERS.copy()
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logging.info(f"  下载尝试 {attempt + 1}/{max_retries}")
                
                response = self.session.get(
                    image_url, 
                    headers=headers, 
                    timeout=30,
                    allow_redirects=True,
                    stream=True
                )
                
                if response.status_code == 200:
                    content = response.content
                    if not content:
                        logging.warning("下载的图片内容为空")
                        if attempt < max_retries - 1:
                            time.sleep(2)
                            continue
                        return None
                    
                    content_type = response.headers.get('Content-Type', '')
                    ext = '.jpg'
                    if 'png' in content_type:
                        ext = '.png'
                    elif 'gif' in content_type:
                        ext = '.gif'
                    elif 'webp' in content_type:
                        ext = '.webp'
                    
                    logging.info(f"  下载成功！文件类型: {content_type}, 大小: {len(content)} 字节")
                    return (content, ext)
                else:
                    logging.warning(f"下载失败，状态码: {response.status_code}")
                    if attempt < max_retries - 1:
                        time.sleep(2 * (attempt + 1))
            
            except Exception as e:
                logging.warning(f"下载异常: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(2)
        
        logging.error("所有下载尝试都失败了")
        return None
    
    def upload_to_agent_bos(self, image_data, file_name):
        """直接上传图片到文心智能体BOS"""
        if not self.cookie:
            return False, "Cookie未加载"
        
        try:
            # 构建表单数据
            files = {
                'type': (None, 'qaImage'),
                'file': (file_name, image_data, 'image/jpeg')
            }
            
            headers = {
                "Accept": "*/*",
                "Cookie": self.cookie,
                "Origin": "https://agents.baidu.com",
                "Referer": "https://agents.baidu.com/lingjing/tools/uploadFileToBos",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
            }
            
            logging.info("发送图片到文心智能体BOS...")
            response = self.session.post(
                self.upload_url,
                files=files,
                headers=headers,
                timeout=60
            )
            
            logging.info(f"响应状态码: {response.status_code}")
            logging.info(f"响应内容: {response.text}")
            
            if response.status_code == 200:
                try:
                    result = response.json()
                    if result.get('errno') == 0 and result.get('data'):
                        image_url = result['data'].get('url')
                        if image_url:
                            logging.info(f"图片上传成功！URL: {image_url}")
                            return True, image_url
                        else:
                            return False, "响应中没有URL"
                    else:
                        return False, result.get('msg', '未知错误')
                except:
                    return False, f"响应解析失败: {response.text}"
            else:
                return False, f"HTTP错误: {response.status_code}"
                
        except Exception as e:
            logging.error(f"上传失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return False, str(e)
    
    def download_and_upload(self, image_url):
        """下载图片并直接上传到文心智能体"""
        # 下载图片
        result = self.download_image(image_url)
        if not result:
            return None
        
        image_data, ext = result
        
        # webp格式转换
        if ext == '.webp':
            logging.info("检测到webp格式，开始转换...")
            image_data, ext = self.convert_image_format(image_data, target_format='JPEG')
        
        # 生成文件名
        file_name = f"image_{int(time.time())}{ext}"
        
        # 上传到文心智能体BOS
        success, result = self.upload_to_agent_bos(image_data, file_name)
        
        if success:
            return result
        else:
            logging.error(f"上传失败: {result}")
            return None


class AgentDataUploader:
    """文心智能体数据上传器"""
    def __init__(self, config):
        agent_config = config.get('agent', {})
        self.cookie_file = agent_config.get('cookie_file')
        self.dataset_id = agent_config.get('dataset_id')
        self.file_id = agent_config.get('file_id')
        self.api_url = agent_config.get('api_url', "https://agents.baidu.com/lingjing/dataset/text/edit/table")
        
        if not self.cookie_file or not os.path.exists(self.cookie_file):
            raise FileNotFoundError(f"Cookie文件不存在: {self.cookie_file}")
        
        self.cookie = self._load_cookie()
    
    def _load_cookie(self):
        """加载Cookie"""
        try:
            with open(self.cookie_file, 'r', encoding='utf-8') as f:
                content = f.read().strip().replace('`', '')
                return content
        except Exception as e:
            logging.error(f"读取Cookie失败: {e}")
            return None
    
    def upload_row(self, title, url, image_url):
        """上传单行数据到文心智能体"""
        if not self.cookie:
            return False, "Cookie未加载"
        
        table_row_detail = [
            {
                "title": "短剧标题",
                "type": 1,
                "content": f"{title}\n" if title else "\n",
                "images": [],
                "imageShield": False
            },
            {
                "title": "短剧简介",
                "type": 1,
                "content": f"{url}\n" if url else "\n",
                "images": [],
                "imageShield": False
            },
            {
                "title": "图片地址",
                "type": 1,
                "content": f"{image_url}\n" if image_url else "\n",
                "images": [],
                "imageShield": False
            }
        ]
        
        payload = {
            "fileId": self.file_id,
            "datasetId": self.dataset_id,
            "tableRowDetail": table_row_detail
        }
        
        headers = {
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "Cookie": self.cookie,
            "Referer": f"https://agents.baidu.com/dataset/detail/paragraph/{self.dataset_id}?fileId={self.file_id}",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36"
        }
        
        try:
            logging.info("发送请求到文心智能体...")
            response = requests.post(
                self.api_url,
                data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),
                headers=headers,
                timeout=60
            )
            
            if response.status_code == 200:
                try:
                    result = response.json()
                    if result.get('errno') == 0:
                        logging.info("上传成功!")
                        return True, result
                    else:
                        return False, result.get('msg', '未知错误')
                except:
                    return True, response.text
            else:
                return False, f"HTTP错误: {response.status_code}"
                
        except Exception as e:
            return False, str(e)


def process_single_row(row_num, ws, excel_config, quark_api, image_uploader, data_uploader):
    """处理单行数据"""
    logging.info(f"\n{'='*60}")
    logging.info(f"处理第 {row_num} 行")
    logging.info(f"{'='*60}")
    
    title_col = excel_config.get('title_column', 1)
    url_col = excel_config.get('url_column', 2)
    image_col = excel_config.get('image_column', 3)
    status_col = 4  # D列：上传状态
    
    # 检查D列是否已标记"已上传"
    status = ws.cell(row=row_num, column=status_col).value
    if status and str(status).strip() == "已上传":
        logging.info(f"D列已标记'已上传'，跳过此行")
        return True, "已上传，跳过"
    
    # 获取标题
    title = ws.cell(row=row_num, column=title_col).value
    logging.info(f"标题: {title}")
    
    # 获取链接列内容
    url_text = ws.cell(row=row_num, column=url_col).value
    logging.info(f"链接内容: {url_text}")
    
    # 检查C列是否已有图片链接
    existing_image = ws.cell(row=row_num, column=image_col).value
    
    # 如果图片链接单元格已有内容，直接使用该链接提交智能体接口
    if existing_image and str(existing_image).strip():
        logging.info(f"图片链接已存在: {existing_image}")
        logging.info(f"直接提交智能体接口...")
        
        # 上传数据到文心智能体
        success, result = data_uploader.upload_row(title, url_text, existing_image)
        if not success:
            logging.error(f"第 {row_num} 行: 上传数据失败: {result}")
            return False, f"上传数据失败: {result}"
        
        # 标记D列为"已上传"
        ws.cell(row=row_num, column=status_col).value = "已上传"
        logging.info(f"第 {row_num} 行处理完成!")
        return True, "成功"
    
    # 如果图片链接单元格为空，需要先上传图片
    logging.info("图片链接为空，需要先上传图片")
    
    # 步骤1: 提取夸克链接
    quark_url = quark_api.extract_quark_url(url_text)
    if not quark_url:
        logging.error(f"第 {row_num} 行: 未找到夸克链接")
        return False, "未找到夸克链接"
    
    # 步骤2: 获取图片URL
    logging.info(f"获取图片URL...")
    image_url, file_name = quark_api.get_first_image_url(quark_url)
    if not image_url:
        logging.error(f"第 {row_num} 行: 无法获取图片URL")
        return False, "无法获取图片URL"
    
    # 步骤3: 直接上传图片到文心智能体BOS
    logging.info(f"下载图片并上传到文心智能体BOS...")
    agent_image_url = image_uploader.download_and_upload(image_url)
    if not agent_image_url:
        logging.error(f"第 {row_num} 行: 上传图片失败")
        return False, "上传图片失败"
    
    # 步骤4: 更新Excel
    logging.info(f"更新Excel...")
    ws.cell(row=row_num, column=image_col).value = agent_image_url
    
    # 步骤5: 上传数据到文心智能体
    logging.info(f"上传数据到文心智能体...")
    success, result = data_uploader.upload_row(title, url_text, agent_image_url)
    if not success:
        logging.error(f"第 {row_num} 行: 上传数据失败: {result}")
        return False, f"上传数据失败: {result}"
    
    # 标记D列为"已上传"
    ws.cell(row=row_num, column=status_col).value = "已上传"
    
    logging.info(f"第 {row_num} 行处理完成!")
    return True, "成功"


def main():
    """主函数"""
    print("=" * 70)
    print("整合上传工具: Excel -> 直接上传到文心智能体")
    print("=" * 70)
    
    try:
        # 加载配置
        config = load_config()
        setup_logging(config)
        
        excel_config = config.get('excel', {})
        excel_path = excel_config.get('input_file')
        
        if not excel_path or not os.path.exists(excel_path):
            logging.error(f"Excel文件不存在: {excel_path}")
            return
        
        # 初始化组件
        logging.info("初始化夸克API...")
        quark_api = QuarkAPI()
        
        logging.info("初始化文心智能体图片上传器...")
        image_uploader = AgentImageUploader(config)
        
        logging.info("初始化文心智能体数据上传器...")
        data_uploader = AgentDataUploader(config)
        
        # 加载Excel
        logging.info(f"加载Excel文件: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb.active
        max_row = ws.max_row
        
        logging.info(f"Excel共有 {max_row} 行数据")
        
        # 统计
        success_count = 0
        fail_count = 0
        
        # 从第2行开始处理
        for row_num in range(2, max_row + 1):
            logging.info(f"\n{'='*60}")
            logging.info(f"开始处理第 {row_num} 行")
            logging.info(f"{'='*60}")
            
            try:
                success, message = process_single_row(
                    row_num, ws, excel_config, quark_api, image_uploader, data_uploader
                )
                
                if success:
                    success_count += 1
                    logging.info(f"第 {row_num} 行成功!")
                else:
                    fail_count += 1
                    logging.error(f"第 {row_num} 行失败: {message}")
                
                # 保存Excel
                save_success = False
                for attempt in range(3):
                    try:
                        wb.save(excel_path)
                        save_success = True
                        logging.info(f"Excel已保存")
                        break
                    except PermissionError:
                        if attempt < 2:
                            logging.warning(f"Excel文件被占用，等待2秒后重试...")
                            time.sleep(2)
                        else:
                            logging.error("Excel文件被占用，无法保存")
                
                # 请求间隔
                delay = config.get('network', {}).get('request_delay', 1)
                if row_num < max_row:
                    logging.info(f"等待 {delay} 秒后继续...")
                    time.sleep(delay)
                
            except Exception as e:
                fail_count += 1
                logging.error(f"第 {row_num} 行异常: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # 最终统计
        print("\n" + "="*70)
        print(f"处理完成! 成功: {success_count}, 失败: {fail_count}")
        print("="*70)
        
    except Exception as e:
        logging.error(f"程序运行出错: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
